from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
import sys
import time
from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import openpyxl
import yaml
from tqdm import tqdm

from openpyxl.styles import Alignment

try:
    from dotenv import dotenv_values
except Exception:
    dotenv_values = None

# OpenAI SDK (Responses API recommended)
from openai import OpenAI  # type: ignore

try:
    # Optional exception classes (vary by SDK version)
    from openai import RateLimitError, APIError, APITimeoutError  # type: ignore
except Exception:  # pragma: no cover
    RateLimitError = APIError = APITimeoutError = Exception  # type: ignore


# -----------------------------
# Utils
# -----------------------------

TAG_RE = re.compile(r"\{[^{}]*\}")
RUBY_RE = re.compile(r"\[([^*\]]+)\*([^\]]+)\]")

# Characters that often appear as extraction noise in this project
NOISE_CHARS = set(["Ц", "Ч"])


def read_yaml(path: str) -> dict:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(path)
    with p.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def normalize_excel_text(s: str) -> str:
    # Excel sometimes contains _x000D_\n sequences
    return (s or "").replace("_x000D_\n", "\n").replace("\r\n", "\n").replace("\r", "\n")


def remove_noise_outside_tags(s: str) -> str:
    # Remove known garbage chars, but keep anything inside { ... } intact.
    if not s:
        return s
    parts = []
    last = 0
    for m in TAG_RE.finditer(s):
        before = s[last:m.start()]
        before = "".join(ch for ch in before if ch not in NOISE_CHARS)
        parts.append(before)
        parts.append(m.group(0))
        last = m.end()
    tail = s[last:]
    tail = "".join(ch for ch in tail if ch not in NOISE_CHARS)
    parts.append(tail)
    return "".join(parts)


def strip_ruby_notation(s: str) -> str:
    # [兄*あに] -> 兄
    return RUBY_RE.sub(lambda m: m.group(1), s or "")


def stable_hash(*items: str) -> str:
    h = sha256()
    for it in items:
        h.update(it.encode("utf-8"))
        h.update(b"\0")
    return h.hexdigest()


def count_newlines(s: str) -> int:
    return (s or "").count("\n")


# -----------------------------
# Cache
# -----------------------------

class SqliteCache:
    def __init__(self, db_path: Path):
        self.db_path = db_path
        self.conn = sqlite3.connect(str(db_path))
        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS cache (k TEXT PRIMARY KEY, v TEXT NOT NULL, created_at INTEGER NOT NULL)"
        )
        self.conn.commit()

    def get(self, k: str) -> Optional[str]:
        cur = self.conn.execute("SELECT v FROM cache WHERE k=?", (k,))
        row = cur.fetchone()
        return row[0] if row else None

    def set(self, k: str, v: str) -> None:
        self.conn.execute(
            "INSERT OR REPLACE INTO cache (k, v, created_at) VALUES (?, ?, ?)",
            (k, v, int(time.time())),
        )
        self.conn.commit()

    def close(self) -> None:
        self.conn.close()


# -----------------------------
# Dictionary bundle
# -----------------------------

@dataclass(frozen=True)
class DictBundle:
    people: Dict[str, dict]
    terms: Dict[str, str]
    relations: dict
    style: dict

    def signature(self) -> str:
        # used for caching. Keep short and deterministic.
        people_sig = stable_hash(json.dumps(self.people, ensure_ascii=False, sort_keys=True))
        terms_sig = stable_hash(json.dumps(self.terms, ensure_ascii=False, sort_keys=True))
        rel_sig = stable_hash(json.dumps(self.relations, ensure_ascii=False, sort_keys=True))
        style_sig = stable_hash(json.dumps(self.style, ensure_ascii=False, sort_keys=True))
        return stable_hash(people_sig, terms_sig, rel_sig, style_sig)


def load_dicts(people_path: str, terms_path: str, relations_path: str | None, style_path: str | None) -> DictBundle:
    people = read_yaml(people_path)
    terms = read_yaml(terms_path)
    relations = read_yaml(relations_path) if relations_path else {}
    style = read_yaml(style_path) if style_path else {}
    return DictBundle(people=people, terms=terms, relations=relations, style=style)


# -----------------------------
# Excel IO
# -----------------------------

def load_master_dialog_rows(xlsx_path: str) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet, List[dict]]:
    wb = openpyxl.load_workbook(xlsx_path)
    if "master_dialog" not in wb.sheetnames:
        raise ValueError("Sheet 'master_dialog' not found")
    ws = wb["master_dialog"]

    # Expect header row at row 1
    headers = [c.value for c in ws[1]]
    expected = ["scenario_dir", "Label", "CharacterName", "ja", "ko"]
    for e in expected:
        if e not in headers:
            raise ValueError(f"Missing column: {e}. Found: {headers}")

    idx = {h: headers.index(h) + 1 for h in headers}

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {
            "rownum": r,
            "scenario_dir": ws.cell(r, idx["scenario_dir"]).value,
            "Label": ws.cell(r, idx["Label"]).value,
            "CharacterName": ws.cell(r, idx["CharacterName"]).value,
            "ja": normalize_excel_text(ws.cell(r, idx["ja"]).value or ""),
            "ko": normalize_excel_text(ws.cell(r, idx["ko"]).value or ""),
        }
        rows.append(row)

    return wb, ws, rows


def set_wrap_text(cell) -> None:
    """Ensure wrap_text so newlines are visible in Excel."""
    a = cell.alignment
    if a is None:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        return
    # keep other properties if possible
    cell.alignment = Alignment(
        horizontal=a.horizontal,
        vertical=a.vertical or "top",
        text_rotation=a.text_rotation,
        wrap_text=True,
        shrink_to_fit=a.shrink_to_fit,
        indent=a.indent,
        relative_indent=a.relative_indent,
        justify_last_line=a.justify_last_line,
        reading_order=a.reading_order,
        text_rotation_189=a.text_rotation_189,
    )


def write_ko(ws, rows: List[dict], ko_map: Dict[int, str], *, wrap_text: bool = True) -> None:
    headers = [c.value for c in ws[1]]
    ko_col = headers.index("ko") + 1
    for row in rows:
        r = row["rownum"]
        if r in ko_map:
            c = ws.cell(r, ko_col)
            c.value = ko_map[r]
            if wrap_text:
                set_wrap_text(c)


def get_ko_col(ws) -> int:
    headers = [c.value for c in ws[1]]
    return headers.index("ko") + 1


def write_ko_group(ws, ko_col: int, group_rows: List[dict], out_lines: List[str], *, wrap_text: bool = True) -> None:
    for r, out in zip(group_rows, out_lines):
        c = ws.cell(r["rownum"], ko_col)
        c.value = out
        if wrap_text:
            set_wrap_text(c)


def atomic_save_workbook(wb: openpyxl.Workbook, out_path: Path) -> None:
    """Safely write xlsx so an interrupted run doesn't corrupt the output."""
    out_path = out_path.resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = out_path.with_suffix(out_path.suffix + ".tmp")
    wb.save(tmp)
    os.replace(tmp, out_path)


# -----------------------------
# Scanning
# -----------------------------

KATAKANA_WORD_RE = re.compile(r"[ァ-ヴー]{3,}")  # rough: >=3 katakana chars
KANJI_TERM_RE = re.compile(r"[一-龥]{2,}")      # rough: >=2 kanji chars


def scan_unknown_speakers(rows: List[dict], people_dict: Dict[str, dict]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for row in rows:
        spk = (row.get("CharacterName") or "").strip()
        if spk and spk.startswith("@") and spk not in people_dict:
            counts[spk] = counts.get(spk, 0) + 1
    return dict(sorted(counts.items(), key=lambda x: (-x[1], x[0])))


def scan_term_candidates(rows: List[dict], terms_dict: Dict[str, str]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for row in rows:
        s = row.get("ja") or ""
        # Ignore tags; only scan outside tags
        s_no_tags = TAG_RE.sub(" ", s)
        for m in KATAKANA_WORD_RE.finditer(s_no_tags):
            w = m.group(0)
            if w not in terms_dict:
                counts[w] = counts.get(w, 0) + 1
        # Also scan kanji terms (optional)
        for m in KANJI_TERM_RE.finditer(s_no_tags):
            w = m.group(0)
            if w not in terms_dict:
                counts[w] = counts.get(w, 0) + 1
    return dict(sorted(counts.items(), key=lambda x: (-x[1], x[0])))


# -----------------------------
# OpenAI translation
# -----------------------------

def load_api_key(env_file: str) -> str:
    p = Path(env_file)
    if not p.exists():
        raise FileNotFoundError(env_file)

    if dotenv_values is not None:
        vals = dotenv_values(env_file)
        key = vals.get("OPENAI_API_KEY")
        if key:
            return key.strip()

    # Fallback: manual parse
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("OPENAI_API_KEY="):
            return line.split("=", 1)[1].strip().strip('"').strip("'")
    raise ValueError("OPENAI_API_KEY not found in env file")


def build_system_instructions(bundle: DictBundle, *, strict_newlines: bool = True) -> str:
    # Keep system prompt concise
    nl_rule = ""
    if strict_newlines:
        nl_rule = (
            "8) 원문에 포함된 줄바꿈(개행)은 개수와 위치를 반드시 유지하라. "
            "출력에 \\n 문자열을 쓰지 말고 실제 줄바꿈을 넣어라.\n"
        )
    return (
        "너는 '칭송받는 자 시리즈: 모노크롬 뫼비우스' 한국어 현지화 번역가다.\n"
        "목표: 게임에 실제로 넣어도 어색하지 않은 자연스러운 번역.\n"
        "규칙:\n"
        "1) 출력은 반드시 JSON 배열. 입력 배열과 길이 동일.\n"
        "2) 각 요소는 문자열이며, 추가 텍스트/설명/주석 금지.\n"
        "3) 원문 안의 { ... } 태그는 문자 단위로 절대 수정/삭제/추가/이동하지 말 것.\n"
        "4) [漢字*ふりがな] 형태는 출력에서 루비를 제거하고 漢字만 남길 것.\n"
        "5) 원문에 섞인 추출 잡음(예: Ц, Ч)은 { ... } 밖에서는 출력에서 제거.\n"
        "6) 영어는 번역하지 말고 그대로 유지.\n"
        "7) 용어/인물/관계 규칙은 아래 사전 섹션을 반드시 준수.\n"
        + nl_rule
    )


def build_dict_section(bundle: DictBundle, *, terms_subset: Optional[Dict[str, str]] = None) -> str:
    # Deterministic formatting for caching stability.
    people_lines = []
    for k in sorted(bundle.people.keys()):
        v = bundle.people[k]
        if isinstance(v, dict):
            ko = v.get("ko") or v.get("name_ko") or v.get("name") or ""
            gender = v.get("gender") or v.get("sex") or ""
        else:
            ko = str(v)
            gender = ""
        if ko:
            people_lines.append(f"{k}={ko}/{gender}".strip("/"))
    use_terms = terms_subset if terms_subset is not None else bundle.terms
    terms_lines = [f"{k}={use_terms[k]}" for k in sorted(use_terms.keys())]

    rel_lines = []
    if bundle.relations:
        rel_lines.append(yaml.safe_dump(bundle.relations, allow_unicode=True, sort_keys=True).strip())
    style_lines = []
    if bundle.style:
        style_lines.append(yaml.safe_dump(bundle.style, allow_unicode=True, sort_keys=True).strip())

    return (
        "### 인물 사전\n" + "\n".join(people_lines) + "\n\n"
        "### 용어 사전\n" + "\n".join(terms_lines) + "\n\n"
        "### 관계/호칭 규칙\n" + ("\n".join(rel_lines) if rel_lines else "(없음)") + "\n\n"
        "### 문체 규칙\n" + ("\n".join(style_lines) if style_lines else "(없음)") + "\n"
    )


def validate_tag_preservation(src_list: List[str], out_list: List[str]) -> Tuple[bool, str]:
    for i, (src, out) in enumerate(zip(src_list, out_list)):
        src_tags = TAG_RE.findall(src)
        out_tags = TAG_RE.findall(out)
        if src_tags != out_tags:
            return False, f"tag_mismatch at index {i}: src_tags={src_tags} out_tags={out_tags}"
    return True, ""


def validate_newline_preservation(src_list: List[str], out_list: List[str]) -> Tuple[bool, str]:
    for i, (src, out) in enumerate(zip(src_list, out_list)):
        if count_newlines(src) != count_newlines(out):
            return (
                False,
                f"newline_mismatch at index {i}: src_nl={count_newlines(src)} out_nl={count_newlines(out)}",
            )
    return True, ""


def translate_batch(
    client: OpenAI,
    model: str,
    bundle: DictBundle,
    src_lines: List[str],
    temperature: float = 0.2,
    max_output_tokens: int = 4000,
    dict_scope: str = "subset",
    strict_newlines: bool = True,
) -> List[str]:
    sys_inst = build_system_instructions(bundle, strict_newlines=strict_newlines)

    # Reduce prompt cost by only sending terms that appear in this batch.
    terms_subset = None
    if dict_scope == "subset":
        joined = "\n".join(src_lines)
        terms_subset = {k: v for k, v in bundle.terms.items() if k and k in joined}
    dict_block = build_dict_section(bundle, terms_subset=terms_subset)

    user_input = {
        "task": "Translate JA->KO for game localization.",
        "notes": "Return JSON array only. Keep newline structure.",
        "dicts": dict_block,
        "lines": src_lines,
    }

    resp = client.responses.create(
        model=model,
        instructions=sys_inst,
        input=json.dumps(user_input, ensure_ascii=False),
        temperature=temperature,
        max_output_tokens=max_output_tokens,
    )
    text = resp.output_text.strip()

    try:
        data = json.loads(text)
    except Exception:
        text2 = text.strip("` \n")
        data = json.loads(text2)

    if not isinstance(data, list) or len(data) != len(src_lines):
        raise ValueError(
            f"Bad output shape. expected list[{len(src_lines)}], got {type(data)} len={len(data) if isinstance(data, list) else 'n/a'}"
        )

    out_lines: List[str] = []
    for s in data:
        if not isinstance(s, str):
            s = str(s)
        s = strip_ruby_notation(s)
        s = remove_noise_outside_tags(s)
        out_lines.append(s)

    ok, msg = validate_tag_preservation(src_lines, out_lines)
    if not ok:
        raise ValueError(msg)

    if strict_newlines:
        ok2, msg2 = validate_newline_preservation(src_lines, out_lines)
        if not ok2:
            raise ValueError(msg2)

    return out_lines


def chunked(seq: List[dict], n: int) -> List[List[dict]]:
    return [seq[i : i + n] for i in range(0, len(seq), n)]


def is_needs_translation(row: dict, *, blanks_only: bool = True, strict_newlines_resume: bool = True) -> bool:
    ko = (row.get("ko") or "")
    ja = (row.get("ja") or "")
    ko_s = ko.strip()
    ja_s = ja.strip()
    if not ja_s:
        return False

    # 기본: ko가 비어있으면 번역
    if blanks_only:
        if not ko_s:
            return True

        # 추가: 원문엔 개행이 있는데 KO에는 개행이 없으면(또는 개행 수 불일치) 재번역
        if strict_newlines_resume and count_newlines(ja) != count_newlines(ko):
            return True

        return False

    # Extended mode: also retranslate if ko equals ja or contains leftover JP.
    if not ko_s or ko_s == ja_s:
        return True
    if strict_newlines_resume and count_newlines(ja) != count_newlines(ko):
        return True
    return bool(re.search(r"[ぁ-んァ-ヴ一-龥]", ko_s))


def is_rate_limit_error(e: Exception) -> bool:
    if isinstance(e, RateLimitError):
        return True
    msg = str(e)
    return "429" in msg or "rate limit" in msg.lower() or "quota" in msg.lower()


def compute_backoff_seconds(e: Exception, attempt: int) -> float:
    retry_after = None
    for attr in ("retry_after", "retry_after_seconds"):
        if hasattr(e, attr):
            try:
                retry_after = float(getattr(e, attr))
            except Exception:
                retry_after = None
    if retry_after and retry_after > 0:
        return min(120.0, retry_after + 0.5)
    return min(90.0, 2.0 ** attempt + 1.0)


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)

    ap_scan = sub.add_parser("scan", help="scan missing speakers/terms in the excel")
    ap_scan.add_argument("--xlsx", required=True)
    ap_scan.add_argument("--dict-people", required=True)
    ap_scan.add_argument("--dict-terms", required=True)
    ap_scan.add_argument("--top", type=int, default=200)

    ap_tr = sub.add_parser("translate", help="translate ja->ko and write out new xlsx")
    ap_tr.add_argument("--xlsx", required=True)
    ap_tr.add_argument("--out", required=True)
    ap_tr.add_argument("--env-file", required=True)
    ap_tr.add_argument("--model", default="gpt-4.1")
    ap_tr.add_argument("--chunk", type=int, default=40)
    ap_tr.add_argument("--dict-people", required=True)
    ap_tr.add_argument("--dict-terms", required=True)
    ap_tr.add_argument("--dict-relations", default=None)
    ap_tr.add_argument("--dict-style", default=None)
    ap_tr.add_argument("--cache-db", default=".cache.sqlite3")
    ap_tr.add_argument("--max-retries", type=int, default=6)
    ap_tr.add_argument(
        "--retranslate-jp",
        action="store_true",
        default=False,
        help="Also translate rows where ko is non-empty but still looks like Japanese.",
    )
    ap_tr.add_argument(
        "--dict-scope",
        choices=["subset", "full"],
        default="subset",
        help="Send only glossary entries relevant to each chunk (subset) or the full glossary.",
    )
    ap_tr.add_argument("--min-delay", type=float, default=0.25, help="Minimum seconds to sleep between API calls.")
    ap_tr.add_argument(
        "--save-every-batch",
        action="store_true",
        default=True,
        help="Checkpoint by saving the output xlsx after each translated chunk (default).",
    )

    # 새 옵션(기본 ON)
    ap_tr.add_argument(
        "--strict-newlines",
        action="store_true",
        default=True,
        help="Force newline count to match source. Mismatch triggers retry (default ON).",
    )
    ap_tr.add_argument(
        "--strict-newlines-resume",
        action="store_true",
        default=True,
        help="When resuming, retranslate rows whose newline count differs (default ON).",
    )
    ap_tr.add_argument(
        "--wrap-text",
        action="store_true",
        default=True,
        help="Set wrap_text=True for ko cells so Excel shows newlines (default ON).",
    )

    args = ap.parse_args()

    if args.cmd == "scan":
        wb, ws, rows = load_master_dialog_rows(args.xlsx)
        people = read_yaml(args.dict_people)
        terms = read_yaml(args.dict_terms)
        unk_spk = scan_unknown_speakers(rows, people)
        cand_terms = scan_term_candidates(rows, terms)

        print("[unknown_speakers]")
        for k, v in list(unk_spk.items())[: args.top]:
            print(f"{k}\t{v}")
        print("\n[term_candidates]")
        for k, v in list(cand_terms.items())[: args.top]:
            print(f"{k}\t{v}")
        return

    if args.cmd == "translate":
        key = load_api_key(args.env_file)
        client = OpenAI(api_key=key)

        bundle = load_dicts(args.dict_people, args.dict_terms, args.dict_relations, args.dict_style)

        out_path = Path(args.out)
        base_xlsx = out_path if out_path.exists() else Path(args.xlsx)
        wb, ws, rows = load_master_dialog_rows(str(base_xlsx))

        blanks_only = not bool(args.retranslate_jp)
        to_tr = [
            r
            for r in rows
            if is_needs_translation(
                r,
                blanks_only=blanks_only,
                strict_newlines_resume=bool(args.strict_newlines_resume),
            )
        ]

        if not to_tr:
            print("Nothing to translate.")
            atomic_save_workbook(wb, out_path)
            return

        cache = SqliteCache(Path(args.cache_db))
        sig = bundle.signature()
        ko_col = get_ko_col(ws)

        pbar = tqdm(total=len(to_tr), desc="Translating", unit="line")

        for group in chunked(to_tr, args.chunk):
            src_lines = [r["ja"] for r in group]
            cache_key = stable_hash(
                "v3",
                args.model,
                args.dict_scope,
                str(bool(args.strict_newlines)),
                sig,
                json.dumps(src_lines, ensure_ascii=False),
            )

            cached = cache.get(cache_key)
            if cached:
                out_lines = json.loads(cached)
                write_ko_group(ws, ko_col, group, out_lines, wrap_text=bool(args.wrap_text))
                if args.save_every_batch:
                    atomic_save_workbook(wb, out_path)
                pbar.update(len(group))
                continue

            last_err: Optional[Exception] = None

            for attempt in range(args.max_retries + 1):
                try:
                    out_lines = translate_batch(
                        client,
                        args.model,
                        bundle,
                        src_lines,
                        dict_scope=args.dict_scope,
                        strict_newlines=bool(args.strict_newlines),
                    )
                    cache.set(cache_key, json.dumps(out_lines, ensure_ascii=False))
                    write_ko_group(ws, ko_col, group, out_lines, wrap_text=bool(args.wrap_text))
                    if args.save_every_batch:
                        atomic_save_workbook(wb, out_path)
                    last_err = None
                    break
                except Exception as e:
                    last_err = e

                    # 429 / quota / rate limit
                    if is_rate_limit_error(e):
                        time.sleep(compute_backoff_seconds(e, attempt))
                        continue

                    # newline mismatch / tag mismatch도 재시도로 처리(조금 더 쉬었다가)
                    msg = str(e).lower()
                    if "newline_mismatch" in msg or "tag_mismatch" in msg:
                        time.sleep(min(15.0, 1.0 + attempt * 2.0))
                        continue

                    # 기타
                    time.sleep(min(20.0, 1.0 + attempt * 1.5))

            if last_err is not None:
                cache.close()
                raise RuntimeError(f"Failed to translate batch starting at row {group[0]['rownum']}: {last_err}") from last_err

            pbar.update(len(group))
            time.sleep(max(0.0, float(args.min_delay)))

        pbar.close()
        atomic_save_workbook(wb, out_path)
        cache.close()
        print(f"Saved: {out_path}")
        return


if __name__ == "__main__":
    main()
